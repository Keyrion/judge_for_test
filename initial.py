import os
import openpyxl

def init():
    ans =[(['']*12)for i in range(12)]*150
    kinds=['']*12
    wb = openpyxl.load_workbook('students_answer.xlsx')
    sht=wb['Sheet1']

    cnt =int(input('输入总小题数:'))
    stunum=0
    for i in range(1,cnt+1):
        kinds[i]=sht.cell(row=3, column=i+1).value

    tmp=sht.cell(row=stunum+4, column=1).value
    while tmp != 'END':
        stunum=stunum+1
        for i in range(1,1+cnt):
            ans[stunum][i] =sht.cell(row=stunum+3,column=i+1).value
        tmp = sht.cell(row=stunum + 4, column=1).value

    return ans,kinds